
import print_fun.printv1 as printv1
from common.read_full_contract_num import get_full
import sys
import os
from http.server import HTTPServer, BaseHTTPRequestHandler, SimpleHTTPRequestHandler
import json
import time
import ntplib
import cgi
import urllib.parse
import urllib.request
import re
from PySide6.QtWidgets import QApplication, QWidget
from PySide6.QtUiTools import QUiLoader
from PySide6.QtCore import QFile, QThread, Signal, Slot, QSize
# 异步需要的函数载入threading
import threading
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

from common.write_contract import write_contract_by

sys.path.append("./common/")
sys.path.append("./print_fun/")
# 载入print_fun

data = {'code': '200'}
host = ('0.0.0.0', 65500)


class Resquest(BaseHTTPRequestHandler):
 ####### POST functions##################
    def export_excel(self, data):
        
        try:
            print('解析后需要打印的data',data)
            # data 内容是这样 [('type', 'exportExcel'), ('header', '[object Object]'), ('list', '[object Object]')] 要解析出header和list
            # [('type', 'exportExcel'), ('header', '{"合同号":"231215-002","单号":"231215-002-5-1WB_0","客户名称":"","拆单":"","客户地址":"","门板厚度":0,"门板颜色":"","门型":"","门型颜色":"","玻璃总价":0,"木箱总价":0,"工艺费总价":0,"反厂费用":0,"备注":"订单备注","name":"name"}'), ('list', '[{"key":1703048079088,"颜色":"默认颜色","宽":1,"高":1,"厚":1,"数量":0,"单价":0,"总价":0,"备注":"备注"}]')]解析出header和list
            # 解析header
            header = json.loads(data[1][1])
            # 解析list
            list = json.loads(data[2][1])
            
            print('解析后header',header)
            print('解析后list',list)

            wb = Workbook()
            ws = wb.active

            # data = json.loads(data['data'])
            # print('解析后data',data['id'])
            # 表格开始

            # 第一行 内容 + style
            ws.append(['诺梵希外购门板' + ' - '+str(header['合同号'])])
            # 第一行字体25，加粗，微软雅黑，合并所有单元格，字体居中,下边框微软蓝色
            ws['A1'].border = Border(bottom=Side(
                border_style='thin', color='FF000000'))
            ws['A1'].font = Font(name='微软雅黑', size=25, bold=True)
            ws['A1'].alignment = Alignment(horizontal='center')
            ws.merge_cells('A1:M1')


            # 第二行 内容 + style 
            ws.append([
                '客户名称',header['客户名称'],'','','','拆单',header['拆单'],'','','客户地址',header['客户地址'],'','',
            ])
            # B~D合并单元格
            ws.merge_cells('B2:E2')
            # G~J合并单元格
            ws.merge_cells('G2:I2')
            # L~M合并单元格
            ws.merge_cells('K2:M2')

            ws.append([
                '门板厚度',header['门板厚度'],'','','','门板颜色',header['门板颜色'],'','','门型',header['门型'],'','',
            ])
            # B~D合并单元格
            ws.merge_cells('B3:E3')
            # G~J合并单元格
            ws.merge_cells('G3:I3')
            # L~M合并单元格
            ws.merge_cells('K3:M3')

            ws.append([
                '门型颜色',header['门型颜色'],'','','','玻璃总价',header['玻璃总价'],'','','木箱总价',header['木箱总价'],'','',
            ])
            # B~D合并单元格
            ws.merge_cells('B4:E4')
            # G~J合并单元格
            ws.merge_cells('G4:I4')
            # L~M合并单元格
            ws.merge_cells('K4:M4')

            ws.append([
                '工艺费总价',header['工艺费总价'],'','','','反厂费用',header['反厂费用'],'','','备注',header['备注'],'','',
            ])
            # B~D合并单元格
            ws.merge_cells('B5:E5')
            # G~J合并单元格
            ws.merge_cells('G5:I5')
            # L~M合并单元格
            ws.merge_cells('K5:M5')




            # 第2,3,4,5行字体10，加粗，微软雅黑，居中，上下左右边框，背景浅灰色
            for i in range(2, 6):
                for j in range(1, 14):
                    ws.cell(i, j).font = Font(
                        name='微软雅黑', size=10, bold=True)
                    ws.cell(i, j).alignment = Alignment(
                        horizontal='center')
                    ws.cell(i, j).border = Border(top=Side(border_style='thin', color='FF000000'), bottom=Side(
                        border_style='thin', color='FF000000'), left=Side(border_style='thin', color='FF000000'), right=Side(border_style='thin', color='FF000000'))
                    ws.cell(i, j).fill = PatternFill(
                        fill_type='solid',
                        # 微软浅蓝色
                        fgColor='D3D3D3')
                    

            # A列整体宽一点
            ws.column_dimensions['A'].width = 15
                
            
            # G列 H列整体宽点
            ws.column_dimensions['G'].width = 20
            ws.column_dimensions['H'].width = 20
            ws.column_dimensions['K'].width = 20
            ws.column_dimensions['L'].width = 0
            ws.column_dimensions['M'].width = 0
            ws.column_dimensions['B'].width = 15



            # 第三行下边距加宽
            ws.row_dimensions[2].height = 20
            ws.row_dimensions[3].height = 20


            # 分割上面区域和下面区域
            ws.append([''])



            # 表头 
            # region
            
            #  ws.append(['编码', '颜色', '宽', '高', '厚', '材积', '数量', '备注']) 背景微软浅蓝色
            ws.append(['编码', '颜色', '宽', '高', '厚', '数量','单价','总价', '备注'])
            # 第一行字体10，加粗，微软雅黑，居中，上下左右边框，背景微软蓝色
            for i in range(1, 10):
                ws.cell(7, i).font = Font(
                    name='微软雅黑', size=10, bold=True)
                ws.cell(7, i).alignment = Alignment(
                    horizontal='center')
                ws.cell(7, i).border = Border(top=Side(border_style='thin', color='FF000000'), bottom=Side(
                    border_style='thin', color='FF000000'), left=Side(border_style='thin', color='FF000000'), right=Side(border_style='thin', color='FF000000'))
                ws.cell(7, i).fill = PatternFill(
                    fill_type='solid',
                    # 微软浅蓝色
                    fgColor='00A4ED')
                
            # 备注这列横跨剩下的
            ws.merge_cells('I7:M7')

                    
            # endregion

            
            


            # 表格内容
            # 从第8行开始写入数据
            # for i in range(0, len(list)):
            #     ws.append([
            #         list[i]['key'],list[i]['颜色'],list[i]['宽'],list[i]['高'],list[i]['厚'],list[i]['材积'],list[i]['数量'],list[i]['备注']
            #     ])
            for i in range(0, len(list)):
                ws.append([
                    str(list[i]['key']),list[i]['颜色'],list[i]['宽'],list[i]['高'],list[i]['厚'],list[i]['数量'],list[i]['单价'],list[i]['总价'],list[i]['备注']
                ])
            # 表格样式
            # 第8行开始字体10，微软雅黑，居中，上下左右边框 行间距稍大
            for i in range(8, len(list)+8):
                for j in range(1, 10):
                    ws.cell(i, j).font = Font(
                        name='微软雅黑', size=10)
                    ws.cell(i, j).alignment = Alignment(
                        horizontal='center')
                    ws.cell(i, j).border = Border(top=Side(border_style='thin', color='FF000000'), bottom=Side(
                        border_style='thin', color='FF000000'), left=Side(border_style='thin', color='FF000000'), right=Side(border_style='thin', color='FF000000'))
                    ws.cell(i, j).alignment = Alignment(
                        vertical='center')
                    ws.row_dimensions[i].height = 20
                    # I列横跨剩余所有格
                    ws.merge_cells('I'+str(i)+':M'+str(i))
                    
            # 表格样式
                    


            # 表格结束
            # wb.save(''+str(header['id'])+'.xlsx')
            # 导出到output目录下
            wb.save('./output/'+str(header['合同号'])+'.xlsx')

        except Exception as e:
            print('导出失败', e)
            window.thread_1.call_back(
                '<font color="red">' + '导出失败,最可能发生这个错误的原因是文件正在被打开,被另外一个进程占用!: ' +
                str(e) + '</font>'
            )
            return {
                'code': 500,
                'status': False,
                'msg': str(e)
            }

        print('导出程序执行')



    
    def do_POST(self):
        data_post = {}
        # 获取post提交的数据
        ctype, pdict = cgi.parse_header(self.headers['content-type'])
        print('ctype:', ctype)
        print('pdict:', pdict)
        dt = self.rfile.read(int(self.headers['content-length']))
        print('dt:', dt)
        # dt现在返回是上面那样，从里面解析出 key 和 value
        pattern = r'name="(.+?)"\r\n\r\n(.+?)\r\n'
        result = re.findall(pattern, dt.decode('utf-8'))
        print('result⚠️:', result)
        # 在textEdit上输出result
        window.thread_1.call_back(
            '<font color="#595">Print cmd : </font>' + str(result)
        )
        # result: [('type', 'exportExcel'), ('header', '[object Object]'), ('list', '[object Object]')] 判断result里面type的值时exportExce还是print
        # 如果是exportExcel,调用exportExcel函数
        if result[0][1] == 'exportExcel':
            print('执行导出 exportExcel',result)
            # 调用exportExcel函数
            self.export_excel(result)
            data_post['code'] = 200
            data_post['status'] = True
            data_post['msg'] = 'exportExcel done'
            self.send_response(200)
            self.send_header("Content-type", "application/json;charset=utf-8")
            self.end_headers()
            self.wfile.write(json.dumps(data_post).encode())
            window.thread_1.call_back(
                '<font color="green">' + '导出信号发送完毕!: ' +
                str(data_post['msg']) + '</font>'
            )
            return
        # 如果是print,调用print函数
        elif result[0][1] == 'print':
            print('print')
            # 调用print函数
            printv1.print_full(
                qr_code=result[1][1],
                text1=result[2][1],
                text2=result[3][1],
                text3=result[4][1]
            )
            data_post['code'] = 200
            data_post['status'] = True
            data_post['msg'] = 'print done'
            self.send_response(200)
            self.send_header("Content-type", "application/json;charset=utf-8")
            self.end_headers()
            self.wfile.write(json.dumps(data_post).encode())
            window.thread_1.call_back(
                '<font color="green">' + '打印信号发送完毕!: ' +
                str(data_post['msg']) + '</font>'
            )
            return
        
        # 如果是test,直接返回成功
        elif result[0][1] == 'test':
            print('test')
            
            data_post['code'] = 200
            data_post['status'] = True
            data_post['msg'] = '导出打印接口连接正常' + printv1.get_printer_name()
            self.send_response(200)
            self.send_header("Content-type", "application/json;charset=utf-8")
            self.end_headers()
            self.wfile.write(json.dumps(data_post).encode())
            window.thread_1.call_back(
                '<font color="green">' + '收到系统测试信号!: ' +
                str(data_post['msg']) + '</font>'
            )
            return


    def do_GET(self):
        self.send_response(200)
        self.send_header('Content-type', 'application/json')
        self.end_headers()
        data['command_loopback'] = self.path
        data['contract_num'] = self.path[5:]
        # 如果command为get,如果合同号符合要求,继续
        command = self.path[0:5]
        content = self.path[5:]
        print('command:', command)
        print('content:', content)
        if content == "":
            content = 'nodata'

    # 获取到命令，开始下一步处理#################
    ################### command get#######################################################################
        if command == '/get/':
            print('command:', command)
            data['code'] = '200'
            data['detail'] = 'success'
            # 发送数据库请求数据，数据库错误应随时中断返回错误数据
            ref_data = get_full(content)

            # 判断查询状态是否成功，成功返回数据，失败返回错误原因
            # print('查询数据结果：',ref_data)
            if ref_data[0] == True:
                # print('数据返回成功d:',ref_data[2])到这数据已经成功获取
                data['result'] = ref_data[2]
                # 根据alpha jobID 开始读取订单详细信息（这里虽然已经成功获取但也应该加错误处理！）
                # print('bug:',data)
                self.wfile.write(json.dumps(data).encode())
            else:
                data['code'] = '501'
                data['detail'] = ref_data[1]
                data['result'] = ref_data[2]
                self.wfile.write(json.dumps(data).encode())
        ########### /command test#######################################################
        elif command == '/test':
            print('This is a test,recived command:', command)
            data['code'] = '200'
            data['detail'] = 'Test success!'
            data['result'] = {}
            # time.sleep(10)
            self.wfile.write(json.dumps(data).encode())
        ########### /unknow command###########################################################
        else:
            print('Unknow command or command error!')
            data['code'] = '500'
            data['detail'] = 'Unknow command or command error!'

            self.wfile.write(json.dumps(data).encode())

        # 如果监测有错误返回命令不能识别,请重试
        # 命令正确，可以调用数据库查询函数处理，并把结果赋值给result，处理过程中错误把code改成错误代码
        # end do get
    def end_headers(self):
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', '*')
        self.send_header('Access-Control-Allow-Headers', '*')
        print('end headers')
        SimpleHTTPRequestHandler.end_headers(self)

    def do_OPTIONS(self):

        self.send_response(200, "ok")
        self.end_headers()

    def do_PUT(self):
        data_put = {}
        ref_data = {}
        self.headers['content-length']
        content_len = int(self.headers['content-length'])
        post_body = self.rfile.read(content_len)
        print("command str :", self.path)
        ref_data = write_contract_by(self.path)

        if ref_data['status'] == True:
            print('写入通过')
            data_put['code'] = 200
            data_put['status'] = True
            data_put['msg'] = 'write done'
            self.send_response(200)
            self.send_header("Content-type", "application/json;charset=utf-8")
            self.end_headers()
            self.wfile.write(json.dumps(data_put).encode())

        else:
            print('写入失败')
            data_put['code'] = 500
            data_put['status'] = False
            data_put['msg'] = ref_data['msg']
            self.send_response(200)
            self.send_header("Content-type", "application/json;charset=utf-8")
            self.end_headers()
            self.wfile.write(json.dumps(data_put).encode())


class Mythread_1(QThread):
    # 可以调用Mythread_1的call_back函数回显数据
    signal_tuple = Signal(tuple)

    def __init__(self):
        super().__init__()

    def run(self):
        server = HTTPServer(host, Resquest)
        self.call_back('<font color="#595">Server start : 诺梵希导出表格,打印贴,alpha家通讯API</font>')
        server.serve_forever()

    def call_back(self, data):
        self.signal_tuple.emit(data)


class MainWindow(QWidget):
    def __init__(self):
        super().__init__()
        qfui = QFile('./start.ui')
        qfui.open(QFile.ReadOnly)
        qfui.close()
        self.ui = QUiLoader().load(qfui)
        # 绑定槽
        self.ui.pushButton.clicked.connect(self.setup_thread_1)
        self.ui.pushButton_2.clicked.connect(self.quit_thread_1)
        # self.ui.textEdit.append("hello world")

    def on_button_click(self):
        # 禁用按键1
        pass
        

    def on_button_click2(self):
        print("Button clicked2!")

    def setup_thread_1(self):
        # 禁用按键
        self.ui.pushButton.setEnabled(False)
        self.thread_1 = Mythread_1()
        self.thread_1.signal_tuple.connect(self.thread_1_finished)
        self.thread_1.start()

    # 销毁thread_1
    def quit_thread_1(self):
        self.thread_1.terminate()
        self.ui.textEdit.append('<font color="#f00"> thread terminated </font>')

    def test(self):
        print('test')

    @Slot(tuple)
    def thread_1_finished(self, item):
        print('接收到子线程1结束信号:', item)
        self.ui.textEdit.append(
            str(item)
        )


if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = MainWindow()
    window.ui.show()
    app.exec()
