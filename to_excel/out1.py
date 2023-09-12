from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side

wb = Workbook()

ws = wb.active

treeData = [["Type", "Leaf Color", "Height"], ["Maple", "Red", 549], [
    "Oak", "Green", 783], ["Pine", "Green", 1204]]

for row in treeData:
    ws.append(row)

# Bold the first row 第一行加粗
ft = Font(bold=True)
for row in ws["A1:C1"]:
    for cell in row:
        cell.font = ft

# 第一行背景色为浅蓝色
for row in ws["A1:C1"]:
    for cell in row:
        cell.fill = PatternFill(fgColor="00CCFF", fill_type="solid")
# C列跨行到F列
ws.merge_cells("C1:F1")

# 第一行高度增加
ws.row_dimensions[1].height = 30

# 第一行右侧实线边框
# for row in ws["A1:C1"]:
#     for cell in row:
#         cell.border = Border(
#             right=Side(style="thin"))

# 整体边框
for row in ws["A1:C4"]:
    for cell in row:
        cell.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"))
        


wb.save("out1.xlsx")
